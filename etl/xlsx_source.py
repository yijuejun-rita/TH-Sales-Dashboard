"""Local-testing data source: reads the workbook via openpyxl and returns the
same {sheet_name: grid} shape that sheets_source.py returns from the live
Google Sheets API, so build_json.py never needs to know which one fed it."""
import openpyxl


def load_grids(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    grids = {}
    for name in wb.sheetnames:
        ws = wb[name]
        grid = []
        for row in ws.iter_rows(values_only=True):
            grid.append(list(row))
        grids[name] = grid
    return grids, list(wb.sheetnames)
